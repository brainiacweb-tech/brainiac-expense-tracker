import os
import csv
import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from typing import List, Dict, Union

# Openpyxl for beautiful native Excel stylesheet exports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Matplotlib integration for advanced GUI plotting
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

# Data Structure types
Transaction = Dict[str, Union[int, str, float]]
CategoryList = List[str]


class BrainiacBudgetTrackerGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Brainiac Budget Tracker Pro")
        self.root.geometry("1150x720")
        self.root.configure(bg="#1E1E2E")
        
        try:
            self.root.tk.call('tk', 'scaling', 1.5)
        except Exception:
            pass

        # Color Palette (Catppuccin Mocha inspired)
        self.bg_color = "#1E1E2E"
        self.card_color = "#252538"
        self.accent_color = "#89B4FA"   # Soft blue
        self.accent_hover = "#B4BEFE"   # Soft lavender
        self.text_primary = "#CDD6F4"
        self.text_secondary = "#A6ADC8"
        self.success_color = "#A6E3A1"  # Soft green
        self.error_color = "#F38BA8"    # Soft red
        self.input_bg = "#313244"
        
        # Core States & Rules
        self.budget_cap = 0.0  # Starts empty as requested
        self.categories: CategoryList = []
        self.transactions: List[Transaction] = []

        # Initialize Enterprise Systems
        self.init_database()
        self.load_data_from_db()
        self.setup_styles()
        self.create_widgets()
        
        # Initial Render Refresh
        self.update_totals()

    def init_database(self):
        """Initializes the SQLite storage layers for long-term tracking next to the script itself."""
        script_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(script_dir, "brainiac_budget.db")
        self.conn = sqlite3.connect(db_path)
        self.cursor = self.conn.cursor()
        
        # Create categories reference table
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        """)
        
        # Create transactions logging table
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                day TEXT NOT NULL,
                category TEXT NOT NULL,
                amount REAL NOT NULL,
                payment_method TEXT NOT NULL,
                description TEXT DEFAULT ''
            )
        """)
        # Ensure description/breakdown column exists (Migration for existing databases)
        try:
            self.cursor.execute("ALTER TABLE transactions ADD COLUMN description TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        self.conn.commit()

        # Seed initial categories if table is completely empty
        self.cursor.execute("SELECT COUNT(*) FROM categories")
        if self.cursor.fetchone()[0] == 0:
            default_cats = ['Food', 'Transport', 'Electricity Bill', 'Water Bill', 'Rent', 'Miscellaneous']
            self.cursor.executemany("INSERT INTO categories (name) VALUES (?)", [(cat,) for cat in default_cats])
            self.conn.commit()

    def load_data_from_db(self):
        """Fetches historical records out of SQL engine into operational application memory."""
        # Load Categories
        self.cursor.execute("SELECT name FROM categories ORDER BY name ASC")
        self.categories = [row[0] for row in self.cursor.fetchall()]

        # Load Transactions
        self.cursor.execute("SELECT id, day, category, amount, payment_method, description FROM transactions")
        self.transactions = []
        for row in self.cursor.fetchall():
            self.transactions.append({
                "id": row[0],
                "day": row[1],
                "category": row[2],
                "amount": row[3],
                "payment_method": row[4],
                "description": row[5] if row[5] is not None else ""
            })

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('default')
        
        # Treeview styling
        style.configure("Treeview",
                        background=self.card_color,
                        foreground=self.text_primary,
                        rowheight=28,
                        fieldbackground=self.card_color,
                        borderwidth=0)
        style.map("Treeview",
                  background=[('selected', self.accent_color)],
                  foreground=[('selected', self.bg_color)])
        
        style.configure("Treeview.Heading",
                        background=self.input_bg,
                        foreground=self.text_primary,
                        font=("Segoe UI", 10, "bold"),
                        borderwidth=0)
        style.map("Treeview.Heading",
                  background=[('active', self.accent_color)],
                  foreground=[('active', self.bg_color)])

        # Combobox styling
        style.configure("TCombobox",
                        fieldbackground=self.input_bg,
                        background=self.accent_color,
                        foreground=self.text_primary,
                        arrowcolor=self.text_primary)

    def create_widgets(self):
        # Header Block with Export button on the far right
        header_frame = tk.Frame(self.root, bg=self.bg_color, pady=10, padx=15)
        header_frame.pack(fill=tk.X)
        
        title_frame = tk.Frame(header_frame, bg=self.bg_color)
        title_frame.pack(side=tk.LEFT)

        title_label = tk.Label(title_frame, text="BRAINIAC BUDGET TRACKER PRO", 
                               font=("Segoe UI", 20, "bold"), fg=self.accent_color, bg=self.bg_color)
        title_label.pack(anchor=tk.W)
        
        subtitle_label = tk.Label(title_frame, text="Data Intelligence & Analytics Dashboard", 
                                  font=("Segoe UI", 10, "italic"), fg=self.text_secondary, bg=self.bg_color)
        subtitle_label.pack(anchor=tk.W)

        # Export button placed prominently in the top header
        self.export_btn = tk.Button(header_frame, text="📤 Export Excel Ledger", font=("Segoe UI", 10, "bold"), bg="#414559", fg=self.text_primary, activebackground=self.accent_color, activeforeground=self.bg_color, bd=0, cursor="hand2", command=self.export_to_excel)
        self.export_btn.pack(side=tk.RIGHT, ipady=6, ipadx=10, pady=5)

        # Split Container
        main_container = tk.Frame(self.root, bg=self.bg_color, padx=15, pady=5)
        main_container.pack(fill=tk.BOTH, expand=True)

        # ==========================================
        # LEFT PANEL: ADD TRANSACTION & UTILITIES
        # ==========================================
        left_panel = tk.Frame(main_container, bg=self.bg_color, width=320)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))

        # Card 1: Add Expense Form
        form_card = tk.LabelFrame(left_panel, text=" Record Expense ", font=("Segoe UI", 11, "bold"), 
                                  fg=self.accent_color, bg=self.card_color, padx=15, pady=10, bd=1, relief=tk.FLAT)
        form_card.pack(fill=tk.X, pady=(0, 10))

        tk.Label(form_card, text="Amount (GH₵):", font=("Segoe UI", 10), fg=self.text_primary, bg=self.card_color).pack(anchor=tk.W, pady=(2, 2))
        self.amount_var = tk.StringVar()
        self.amount_entry = tk.Entry(form_card, textvariable=self.amount_var, font=("Segoe UI", 11), bg=self.input_bg, fg=self.text_primary, bd=0, insertbackground=self.text_primary)
        self.amount_entry.pack(fill=tk.X, ipady=4, pady=(0, 8))

        tk.Label(form_card, text="Category:", font=("Segoe UI", 10), fg=self.text_primary, bg=self.card_color).pack(anchor=tk.W, pady=(2, 2))
        self.category_var = tk.StringVar()
        self.category_dropdown = ttk.Combobox(form_card, textvariable=self.category_var, state="readonly")
        self.category_dropdown['values'] = self.categories
        self.category_dropdown.pack(fill=tk.X, ipady=4, pady=(0, 8))
        if self.categories: self.category_dropdown.current(0)

        tk.Label(form_card, text="Description / Breakdown:", font=("Segoe UI", 10), fg=self.text_primary, bg=self.card_color).pack(anchor=tk.W, pady=(2, 2))
        self.description_var = tk.StringVar()
        self.description_entry = tk.Entry(form_card, textvariable=self.description_var, font=("Segoe UI", 11), bg=self.input_bg, fg=self.text_primary, bd=0, insertbackground=self.text_primary)
        self.description_entry.pack(fill=tk.X, ipady=4, pady=(0, 8))

        tk.Label(form_card, text="Date / Day:", font=("Segoe UI", 10), fg=self.text_primary, bg=self.card_color).pack(anchor=tk.W, pady=(2, 2))
        self.day_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.day_entry = tk.Entry(form_card, textvariable=self.day_var, font=("Segoe UI", 11), bg=self.input_bg, fg=self.text_primary, bd=0, insertbackground=self.text_primary)
        self.day_entry.pack(fill=tk.X, ipady=4, pady=(0, 8))

        tk.Label(form_card, text="Payment Method:", font=("Segoe UI", 10), fg=self.text_primary, bg=self.card_color).pack(anchor=tk.W, pady=(2, 2))
        self.payment_var = tk.StringVar(value="Cash")
        self.payment_entry = tk.Entry(form_card, textvariable=self.payment_var, font=("Segoe UI", 11), bg=self.input_bg, fg=self.text_primary, bd=0, insertbackground=self.text_primary)
        self.payment_entry.pack(fill=tk.X, ipady=4, pady=(0, 12))

        self.add_btn = tk.Button(form_card, text="Add Expense", font=("Segoe UI", 11, "bold"), bg=self.accent_color, fg=self.bg_color, activebackground=self.accent_hover, activeforeground=self.bg_color, bd=0, cursor="hand2", command=self.add_expense)
        self.add_btn.pack(fill=tk.X, ipady=6)

        # Card 2: Create Custom Category
        category_card = tk.LabelFrame(left_panel, text=" Category Strategy ", font=("Segoe UI", 11, "bold"), fg=self.accent_color, bg=self.card_color, padx=15, pady=10, bd=1, relief=tk.FLAT)
        category_card.pack(fill=tk.X, pady=(0, 10))

        self.new_cat_var = tk.StringVar()
        self.new_cat_entry = tk.Entry(category_card, textvariable=self.new_cat_var, font=("Segoe UI", 11), bg=self.input_bg, fg=self.text_primary, bd=0, insertbackground=self.text_primary)
        self.new_cat_entry.pack(fill=tk.X, ipady=4, pady=(0, 8))

        self.cat_btn = tk.Button(category_card, text="Create Category", font=("Segoe UI", 10, "bold"), bg=self.input_bg, fg=self.text_primary, activebackground=self.accent_color, activeforeground=self.bg_color, bd=0, cursor="hand2", command=self.create_category)
        self.cat_btn.pack(fill=tk.X, ipady=5, pady=(0, 10))

        self.clear_btn = tk.Button(category_card, text="🗑️ Reset / Clear All Data", font=("Segoe UI", 10, "bold"), bg="#E05B5B", fg="#FFFFFF", activebackground="#F38BA8", activeforeground="#FFFFFF", bd=0, cursor="hand2", command=self.clear_database_data)
        self.clear_btn.pack(fill=tk.X, ipady=5)

        # ==========================================
        # RIGHT PANEL: LOGS & LIVE VISUAL ANALYTICS
        # ==========================================
        right_panel = tk.Frame(main_container, bg=self.bg_color)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Top half: Dynamic Metrics Card with live Target Input
        totals_card = tk.Frame(right_panel, bg=self.card_color, padx=15, pady=15)
        totals_card.pack(fill=tk.X, pady=(0, 10))

        self.total_label = tk.Label(totals_card, text="Total Expenses: GH₵0.00", font=("Segoe UI", 18, "bold"), fg=self.success_color, bg=self.card_color)
        self.total_label.pack(side=tk.LEFT)
        
        target_frame = tk.Frame(totals_card, bg=self.card_color)
        target_frame.pack(side=tk.RIGHT)

        tk.Label(target_frame, text="Target Budget (GH₵):", font=("Segoe UI", 11, "bold"), fg=self.text_secondary, bg=self.card_color).pack(side=tk.LEFT, padx=(10, 5))

        self.budget_var = tk.StringVar(value="")  # Starts completely empty for user to input
        self.budget_entry = tk.Entry(target_frame, textvariable=self.budget_var, font=("Segoe UI", 11, "bold"), bg=self.input_bg, fg=self.accent_color, width=10, bd=0, insertbackground=self.text_primary, justify=tk.CENTER)
        self.budget_entry.pack(side=tk.LEFT, padx=5, ipady=2)
        self.budget_entry.bind("<KeyRelease>", lambda e: self.update_budget_limit())

        self.status_label = tk.Label(target_frame, text="🟢 Stable", font=("Segoe UI", 11, "bold"), fg=self.success_color, bg=self.card_color)
        self.status_label.pack(side=tk.LEFT, padx=(15, 0))

        # Sub-container for historical records and visual matplotlib charts
        data_container = tk.Frame(right_panel, bg=self.bg_color)
        data_container.pack(fill=tk.BOTH, expand=True)

        # Left Column inside Data Container: Interactive logs table
        history_card = tk.LabelFrame(data_container, text=" Transaction Logs ", font=("Segoe UI", 11, "bold"), fg=self.accent_color, bg=self.card_color, padx=10, pady=10, bd=1, relief=tk.FLAT)
        history_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        scrollbar = ttk.Scrollbar(history_card)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree = ttk.Treeview(history_card, columns=("Date", "Category", "Description", "Amount", "Payment"), show="headings", yscrollcommand=scrollbar.set)
        self.tree.heading("Date", text="Date/Day")
        self.tree.heading("Category", text="Category")
        self.tree.heading("Description", text="Description/Breakdown")
        self.tree.heading("Amount", text="Amount (GH₵)")
        self.tree.heading("Payment", text="Payment Method")

        self.tree.column("Date", width=95, anchor=tk.CENTER)
        self.tree.column("Category", width=105, anchor=tk.W)
        self.tree.column("Description", width=130, anchor=tk.W)
        self.tree.column("Amount", width=85, anchor=tk.E)
        self.tree.column("Payment", width=95, anchor=tk.CENTER)

        self.tree.pack(fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.tree.yview)
        self.tree.bind("<Delete>", self.delete_selected_expense)

        # Repopulate visual records onto tree view from database rows loaded at initial startup
        for tx in self.transactions:
            self.tree.insert("", tk.END, iid=str(tx["id"]), values=(tx["day"], tx["category"], tx.get("description", ""), f"GH₵{tx['amount']:.2f}", tx["payment_method"]))

        # Right Column inside Data Container: Modern Dynamic Matplotlib Pie Chart Panel
        self.chart_card = tk.LabelFrame(data_container, text=" Real-Time Categorical Share ", font=("Segoe UI", 11, "bold"), fg=self.accent_color, bg=self.card_color, padx=5, pady=5, bd=1, relief=tk.FLAT, width=360)
        self.chart_card.pack(side=tk.RIGHT, fill=tk.BOTH)
        self.chart_card.pack_propagate(False)

        # Setup Embedded Plotting Canvas
        self.fig = Figure(figsize=(3.5, 3.5), dpi=100, facecolor=self.card_color)
        self.ax = self.fig.add_subplot(111)
        self.ax.set_facecolor(self.card_color)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.chart_card)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)



    def create_category(self):
        new_category = self.new_cat_var.get().strip().title()
        
        if not new_category:
            messagebox.showwarning("Empty Name", "Please enter a valid category name.")
            return
        if new_category in self.categories:
            messagebox.showwarning("Duplicate Category", f"'{new_category}' category already exists.")
            return

        # Write to Database
        try:
            self.cursor.execute("INSERT INTO categories (name) VALUES (?)", (new_category,))
            self.conn.commit()
            
            self.categories.append(new_category)
            self.categories.sort()
            self.category_dropdown['values'] = self.categories
            self.category_dropdown.set(new_category)
            self.new_cat_var.set("")
            messagebox.showinfo("Success", f"Category '{new_category}' created and saved to db!")
            self.update_totals()
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Could not save category: {e}")

    def add_expense(self):
        try:
            amount_str = self.amount_var.get().strip()
            if not amount_str: raise ValueError()
            amount = float(amount_str)
            if amount < 0: raise ValueError()
        except ValueError:
            messagebox.showerror("Invalid Amount", "Please enter a positive numeric value for the amount.")
            return

        category = self.category_var.get().strip()
        if not category:
            messagebox.showerror("Invalid Category", "Please select or create a category first.")
            return

        day = self.day_var.get().strip()
        if not day: day = datetime.now().strftime("%Y-%m-%d")
        payment = self.payment_var.get().strip()
        if not payment: payment = "Cash"
        description = self.description_var.get().strip()

        # Database Insertion Action
        try:
            self.cursor.execute("""
                INSERT INTO transactions (day, category, amount, payment_method, description)
                VALUES (?, ?, ?, ?, ?)
            """, (day, category, amount, payment, description))
            self.conn.commit()
            generated_id = self.cursor.lastrowid
            
            # Update Local App UI Memory state tracking
            transaction: Transaction = {
                "id": generated_id,
                "category": category,
                "amount": amount,
                "day": day,
                "payment_method": payment,
                "description": description
            }
            self.transactions.append(transaction)

            # Map dynamically directly onto the tree viewer
            self.tree.insert("", tk.END, iid=str(generated_id), values=(day, category, description, f"GH₵{amount:.2f}", payment))
            self.amount_var.set("")
            self.description_var.set("")
            
            self.update_totals()
        except sqlite3.Error as e:
            messagebox.showerror("Database Transaction Error", f"Failed recording execution run parameters: {e}")

    def delete_selected_expense(self, event):
        selected_items = self.tree.selection()
        if not selected_items: return

        if not messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete the selected transaction(s)? This will wipe them from persistent storage."):
            return

        for item in selected_items:
            db_id = int(item)
            try:
                # Remove instantly from DB engine
                self.cursor.execute("DELETE FROM transactions WHERE id = ?", (db_id,))
                self.conn.commit()
                
                # Strip out of UI tracking array lists
                self.transactions = [tx for tx in self.transactions if tx["id"] != db_id]
                self.tree.delete(item)
            except sqlite3.Error as e:
                messagebox.showerror("SQL Wipe Exception", f"Could not clear record key entries safely: {e}")
                
        self.update_totals()

    def update_budget_limit(self):
        """Updates threshold limits dynamically from input field focus actions."""
        try:
            val_str = self.budget_var.get().strip()
            if not val_str:
                self.budget_cap = 0.0
                self.update_totals()
                return
            val = float(val_str)
            if val >= 0:
                self.budget_cap = val
                self.update_totals()
        except ValueError:
            # Let the user keep typing instead of resetting their cursor/field
            pass

    def update_totals(self):
        """Processes analytics metrics dynamically and redraws charts in real-time."""
        total = sum(float(tx["amount"]) for tx in self.transactions)
        self.total_label.config(text=f"Total Expenses: GH₵{total:.2f}")

        # Budget cap monitoring check system rules
        if self.budget_cap <= 0:
            self.total_label.config(fg=self.success_color)
            self.status_label.config(text="🟢 Stable", fg=self.success_color)
        elif total > self.budget_cap:
            self.total_label.config(fg=self.error_color)
            self.status_label.config(text="⚠️ Over Budget", fg=self.error_color)
        else:
            self.total_label.config(fg=self.success_color)
            self.status_label.config(text="🟢 Stable", fg=self.success_color)

        # Aggregate categorical datasets for matplotlib mapping pipeline
        category_totals: Dict[str, float] = {}
        for tx in self.transactions:
            cat = str(tx["category"])
            category_totals[cat] = category_totals.get(cat, 0.0) + float(tx["amount"])

        # Strip active groups to avoid throwing exceptions with empty slices on plots
        active_slices = {k: v for k, v in category_totals.items() if v > 0}

        # Redraw chart safely without freezing application main threads
        self.ax.clear()
        if active_slices:
            labels = list(active_slices.keys())
            sizes = list(active_slices.values())
            
            # Clean modern color map for rendering
            colors = ['#89B4FA', '#A6E3A1', '#F38BA8', '#FAB387', '#CBA6F7', '#F9E2AF', '#94E2D5']
            
            self.ax.pie(
                sizes, 
                labels=labels, 
                autopct='%1.1f%%', 
                startangle=140, 
                colors=colors[:len(labels)],
                textprops={'color': self.text_primary, 'fontsize': 8}
            )
            self.fig.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.1)
        else:
            self.ax.text(0.5, 0.5, "No Expense Logs Registered", color=self.text_secondary, ha='center', va='center')
            self.ax.axis('off')
            
        self.canvas.draw()

    def export_to_excel(self):
        """Saves historical records out to a beautifully styled native Excel (.xlsx) workbook."""
        if not self.transactions:
            messagebox.showwarning("No Data", "There are no transactions available to export.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbooks", "*.xlsx"), ("All files", "*.*")],
            title="Export Premium Ledger"
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ledger"

            # Colors and Fills
            header_fill = PatternFill(start_color="252538", end_color="252538", fill_type="solid")
            zebra_fill = PatternFill(start_color="F4F5F8", end_color="F4F5F8", fill_type="solid")
            total_fill = PatternFill(start_color="EAEBEF", end_color="EAEBEF", fill_type="solid")
            
            # Fonts
            font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_data = Font(name="Segoe UI", size=10, color="2C2C2C")
            font_total = Font(name="Segoe UI", size=11, bold=True, color="1E1E2E")
            
            # Borders
            thin_border_side = Side(border_style="thin", color="D1D2D6")
            border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            double_bottom = Border(top=Side(style="thin", color="1E1E2E"), bottom=Side(style="double", color="1E1E2E"))

            # Title Banner (Rows 1-2 merged)
            ws.merge_cells("A1:F2")
            ws["A1"] = "BRAINIAC FINANCIAL LEDGER"
            ws["A1"].font = font_title
            ws["A1"].fill = header_fill
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

            # Table Headers (Row 4)
            headers = ["Transaction ID", "Date", "Category", "Description", "Amount", "Payment Method"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col_idx, value=header)
                cell.font = font_header
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data Rows (starting at Row 5)
            start_row = 5
            for row_idx, tx in enumerate(self.transactions, start=start_row):
                row_data = [
                    tx["id"],
                    tx["day"],
                    tx["category"],
                    tx.get("description", ""),
                    float(tx["amount"]),
                    tx["payment_method"]
                ]
                
                # Apply alternate zebra shading
                is_even = (row_idx % 2 == 0)
                
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_data
                    cell.border = border_data
                    
                    if is_even:
                        cell.fill = zebra_fill
                        
                    # Alignments & Number Formats
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx == 2:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx == 3:
                        cell.alignment = Alignment(horizontal="left")
                    elif col_idx == 4:
                        cell.alignment = Alignment(horizontal="left")
                    elif col_idx == 5:
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = '"GH₵"#,##0.00'
                    elif col_idx == 6:
                        cell.alignment = Alignment(horizontal="center")

            # Calculations & Grand Totals Row (2 rows below data)
            last_row = start_row + len(self.transactions) - 1
            total_row = last_row + 2
            
            # Merge columns A-D for "Grand Total" label
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
            total_label_cell = ws.cell(row=total_row, column=1, value="Grand Total Expenses:")
            total_label_cell.font = font_total
            total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # Excel SUM Formula
            total_val_cell = ws.cell(row=total_row, column=5, value=f"=SUM(E{start_row}:E{last_row})")
            total_val_cell.font = font_total
            total_val_cell.alignment = Alignment(horizontal="right", vertical="center")
            total_val_cell.border = double_bottom
            total_val_cell.fill = total_fill
            total_val_cell.number_format = '"GH₵"#,##0.00'

            # Auto-adjust column widths dynamically to prevent ### and cut-offs
            ws.row_dimensions[1].height = 20
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[4].height = 28
            
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                
                # Check lengths only from Row 4 downwards (ignore A1 title merge length)
                for cell in col[3:]:
                    if cell.value:
                        # Add extra length buffer for currency prefix formatting
                        cell_len = len(str(cell.value))
                        if col_letter == 'E':
                            cell_len += 6  # Pad for 'GH₵' prefix and '.00' decimals
                        max_len = max(max_len, cell_len)
                
                # Set width with safety margin
                ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

            # Save File
            wb.save(file_path)
            messagebox.showinfo("Export Success", f"Financial ledger exported as a premium Excel sheet to:\n{os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("Export Exception", f"An unforeseen error interrupted data writes:\n{e}")

    def clear_database_data(self):
        """Clears all logged transactions from SQLite database and operational memory."""
        if messagebox.askyesno("Reset Tracker", "Are you sure you want to permanently clear all logged transactions? This will wipe the database."):
            try:
                self.cursor.execute("DELETE FROM transactions")
                self.conn.commit()
                self.transactions = []
                
                # Clear UI Treeview log items
                for item in self.tree.get_children():
                    self.tree.delete(item)
                
                self.update_totals()
                messagebox.showinfo("Success", "All transactions cleared successfully!")
            except sqlite3.Error as e:
                messagebox.showerror("Database Reset Error", f"Failed clearing SQLite rows: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = BrainiacBudgetTrackerGUI(root)
    root.mainloop()